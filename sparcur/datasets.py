import io
import csv
import copy
from types import GeneratorType
from itertools import chain
from collections import Counter, defaultdict
#import openpyxl  # import time hog
import augpathlib as aug
from xlsx2csv import Xlsx2csv, SheetNotFoundException, InvalidXlsxFileException
from terminaltables import AsciiTable
from pyontutils.utils import python_identifier
from . import schemas as sc
from . import raw_json as rj
from . import exceptions as exc
from . import normalization as nml
from .core import log, logd, HasErrors
from .paths import Path, BlackfynnCache
from .utils import is_list_or_tuple


# FIXME review this :/ it is not a good implementation at all
def tos(hrm):
    if isinstance(hrm, GeneratorType):
        t = tuple(hrm)
        if not t:
            return
        elif len(t) == 1:
            return t[0]
        else:
            return t
    else:
        return hrm


def to_string_and_then_python_identifier(thing):
    """ handl cases where a header value isn't a string """
    return python_identifier(str(thing))


hasSchema = sc.HasSchema()
@hasSchema.mark
class Header:
    """ generic header normalization for python """
    # FIXME this is a really a pipeline stage
    def __init__(self, first_row_or_column, normalize=True, alt=False, normalize_first_cell=False):
        self.pipeline_start = first_row_or_column
        self._normalize = normalize
        self._alt = alt
        self._normalize_first_cell = normalize or normalize_first_cell

    @property
    def normalized(self):
        if not hasattr(self, '_cache_normalized'):
            orig_header = self.pipeline_start
            header = []
            prefix = 'TEMPA' if self._alt else 'TEMPH' # prevent collision
            for i, c in enumerate(orig_header):
                if c:
                    c = to_string_and_then_python_identifier(c)
                    c = nml.NormHeader(c)

                if not c:
                    c = f'{prefix}_{i}'

                if c in header:
                    c = c + f'_{prefix}_{i}'

                header.append(c)

            self._cache_normalized = header

        return self._cache_normalized

    @hasSchema(sc.HeaderSchema)
    def data(self):
        return self.normalized

    @property
    def _lut(self):
        return {n:h for n, h in zip(self.data, self.pipeline_start)}

    def lut(self, **renames):
        if self._normalize:
            return self._lut_normalize(renames)
        else:
            return self._lut_original(renames)

    def _lut_normalize(self, renames):
        """ do any renames beyond the specific normalization """
        return {renames[n] if n in renames else n:h
                for n, h in zip(self.data, self.pipeline_start)}

    def _lut_original(self, renames):
        original = self.pipeline_start
        normalized = self.normalized
        if len(set(original)) != len(original):
            dupes = [v for v, c in Counter(original).most_common() if c > 1]
            msg = f'Original header is not unique.\nDuplicate entries: {dupes}'
            if None in dupes:
                msg += '\nThere may be multiple records without a primary key.'
            raise exc.MalformedHeaderError(msg)

        if self._normalize_first_cell:
            norm_first = (*normalized[:1], *original[1:])
            return {renames[n] if n in renames else n:h
                    for n, h in zip(norm_first, original)}
        else:
            return {renames[n] if n in renames else n:h
                    for n, h in zip(original, original)}


class DatasetMetadata(HasErrors):

    def __init__(self, path, template_schema_version=None):
        self._path = path
        self.cache = path.cache
        self.name = path.name
        # XXX transitive updated is not entirely necessary anymore since the
        # dataset updated date is bumped on any transitive modification
        self.updated_cache_transitive = path.updated_cache_transitive
        self.template_schema_version = template_schema_version

    @property
    def data(self):
        if self.cache is not None:
            cmeta = self.cache.meta
            return dict(
                id=self.cache.id,
                meta=dict(
                    folder_name=self.name,
                    uri_human=self.cache.uri_human,
                    uri_api=self.cache.uri_api,
                    timestamp_created=cmeta.created,
                    timestamp_updated=cmeta.updated,
                    # FIXME vs the dataset updated time now touched for transitive updates
                    timestamp_updated_contents=self.updated_cache_transitive(),
                    sparse=self.cache.is_sparse(),
                        ))
        else:
            raise NotImplementedError('use CacheL instead')
            path = self._path
            meta = path.meta
            #uri = quote(path.resolve().as_uri(), safe=':/')
            uri = path.resolve().as_uri()
            return dict(id=meta.id,
                        meta=dict(folder_name=path.name,
                                  uri_human=uri,
                                  uri_api=uri,
                                  timestamp_created=None,  # cannot get birthdate on all file systems
                                  timestamp_updated=meta.updated,
                        ))


class DatasetStructure:
    sections = None  # NOTE assigned in at end of file
    rglobs = 'manifest',
    default_glob = 'glob'
    max_childs = None  # 40 was used when we didn't have sparse pulls
    rate = 8  # set by Integrator from cli
    _refresh_on_missing = True
    _classes = {}

    def _newimpl(cls, path, *args, **kwargs):
        apc = path.__class__._AugmentedPath__abstractpath
        if apc not in cls._classes:
            # keep classes 1:1 with the types of paths so that equality works correctly
            # mro issues with specialized paths instead of the base
            newcls = type(cls.__name__ + apc.__name__, (cls, apc),
                          #cls.__dict__,
                          dict(__new__=apc.__new__,
                               __repr__=apc.__repr__,
                               )
                          )
            newcls._bind_flavours(pos_helpers=(HasErrors,), win_helpers=(HasErrors,))
            cls._classes[apc] = newcls

        nc = cls._classes[apc]
        return nc(path)

    def __new__(cls, path, *args, **kwargs):
        return cls._newimpl(cls, path)

    _renew = __new__

    def __new__(cls, path, *args, **kwargs):
        DatasetStructure._setup()
        DatasetStructure.__new__ = DatasetStructure._renew

        return cls._newimpl(cls, path)

    @staticmethod
    def _setup():
        import requests
        DatasetStructure._requests = requests

    @property
    def children(self):
        # for whatever reason the way pathlib constructs children
        # skips calling init, iirc because they use object.__new__
        # or something

        for child in super().children:
            child.__init__()
            yield child

    @classmethod
    def _bind_sections(cls):
        for section, _sec_class in cls.sections.items():
            @property
            def sec(self, sec_class=_sec_class, name='_cache_' + _sec_class.__name__):
                if not hasattr(self, name):
                    mp = [p for p in self.meta_paths if isinstance(p, sec_class)]
                    if len(mp) == 1:
                        setattr(self, name, mp[0])
                    elif not mp:
                        setattr(self, name, None)
                    else:
                        setattr(self, name, mp)

                return getattr(self, name)

            setattr(cls, section, sec)

    def files_last_updated(self):
        # FIXME TODO can we get this more efficiently during pull?
        all_updated = [c.cache.meta.updated for c in
                       chain((self,), self.rchildren) if c.is_file()]
        if all_updated:
            return max(all_updated)

    @property
    def counts(self):
        if not hasattr(self, '_counts'):
            size = 0
            dirs = 0
            files = 0
            need_meta = []
            if not self.is_dir():
                gen = self,

            else:
                gen = self.rchildren

            for c in gen:
                if c.is_dir():
                    dirs += 1
                    maybe_size = 0  # the remote size of a folder is zero for bf
                else:
                    files += 1  # testing for broken symlinks is hard
                    try:
                        if c.is_broken_symlink():
                            maybe_size = (aug.PathMeta
                                          .from_symlink(c)
                                          .size)
                        else:
                            _meta = aug.PathMeta.from_xattrs(c.xattrs(),
                                                             prefix='bf')
                            if _meta.size is not None:
                                maybe_size = _meta.size
                            elif c.is_file():  # running an export on a local dataset
                                maybe_size = c.size

                        #maybe_size = c.cache.meta.size
                    except AttributeError as e:
                        log.error(f'no cache or no meta for {c}\n{e}')
                        continue

                    if maybe_size is None:
                        need_meta.append(c)
                    else:
                        size += maybe_size

            if need_meta and self._refresh_on_missing:
                msg = "We don't do this anymore. Fetch everything first."
                raise exc.NetworkSandboxError(msg)

            self._counts = dict(size=aug.FileSize(size), dirs=dirs, files=files)

        return self._counts

    @property
    def total_size(self):
        """ total remote size, trigger retrieval of all remote metadata """
        return self.counts['size']

    @property
    def total_dirs(self):
        return self.counts['dirs']

    @property
    def total_files(self):
        return self.counts['files']

    @property
    def total_paths(self):
        return self.counts['dirs'] + self.counts['files']

    @property
    def dataset_root(self):
        # FIXME this will find the first dataset description file at any depth!
        # this is incorrect behavior!
        """ The folder where the dataset_description file is located.
            Sometimes there is an intervening folder. """
        if self.cache.is_dataset:
            def check_fordd(paths, level=0, stop=3):
                if not paths:  # apparently the empty case recurses forever
                    return

                if (self.max_childs is not None and
                    self.max_childs > 0 and
                    len(paths) > self.max_childs):
                    log.warning(f'Not globing in a folder with > {self.max_childs} children! '
                                f'{self.as_posix()!r}')
                    return
                dd_paths_all = []
                children = []
                for path in paths:
                    dd_paths = list(path.glob('[Dd]ataset_description*.*'))
                    if dd_paths:
                        dd_paths_all.extend(dd_paths)
                    elif not dd_paths_all:
                        children.extend([p for p in path.children if p.is_dir()])

                if dd_paths_all:
                    return dd_paths_all
                else:
                    return check_fordd(children, level + 1)

            dd_paths = check_fordd((self,))

            if not dd_paths:
                #log.warning(f'No bids root for {self.name} {self.id}')  # logging in a property -> logspam
                return

            elif len(dd_paths) > 1:
                #log.warning(f'More than one submission for {self.name} {self.id} {dd_paths}')
                pass

            return dd_paths[0].parent  # FIXME choose shorter version? if there are multiple?

        elif self.parent:  # organization has no parent
            return self.parent.dataset_root

    def _abstracted_paths(self, name_prefix, glob_type=None, fetch=True, sandbox=False):
        """ A bottom up search for the closest file in the parent directory.
            For datasets, if the bids root and path do not match, use the bids root.
            In the future this needs to be normalized because the extra code required
            for dealing with the intervening node is quite annoying to maintain.
        """
        if glob_type is None:
            glob_type = self.default_glob

        path = self
        cache = self.cache
        if cache and cache == cache.anchor:  # organization
            for child in self.children:
                yield from child._abstracted_paths(name_prefix,
                                                   glob_type,
                                                   fetch,
                                                   sandbox,)
            return
        elif (self.cache and
              self.cache.is_dataset and
              self.dataset_root is not None and
              self.dataset_root != self):
            path = self.dataset_root

        first = name_prefix[0]
        cased_np = '[' + first.upper() + first + ']' + name_prefix[1:]  # FIXME warn and normalize
        glob = getattr(path, glob_type)
        gen = glob(cased_np + '*.*')

        try:
            path_class = self.sections[name_prefix]
            path = next(gen)
            for path in chain((path,), gen):
                path = path_class(path)
                if path.is_broken_symlink():
                    if not fetch:
                        yield path
                        continue
                    elif sandbox:
                        msg = ('A required metadata path was not fetched '
                               f'in advance. {path}')
                        raise exc.NetworkSandboxError(msg)

                    log.info(f'fetching unretrieved metadata path {path.as_posix()!r}'
                             '\nFIXME batch these using async in cli export ...')
                    try:
                        path.cache.fetch(size_limit_mb=path.cache.meta.size.mb + 1)
                    except self._requests.exceptions.ConnectionError as e:
                        raise exc.NetworkFailedForPathError(path) from e

                name = path.name
                if (path.is_file() and
                    (name.startswith('.~lock.') and name.endswith('#')) and
                    path.cache.meta.id is None):
                    # FIXME abstract to handle arbitrary path name regex patterns
                    log.warning(f'illegal untracked file detect {path}')
                    continue

                if path.suffix in path.stem:
                    msg = f'path has duplicate suffix {path.as_posix()!r}'
                    self.addError(msg,
                                  blame='submission',
                                  path=path)
                    logd.error(msg)

                if name[0].isupper():  # FIXME should probably be checking name.lower() != name
                    msg = f'path has bad casing {path.as_posix()!r}'
                    if self.addError(msg,
                                     blame='submission',
                                     path=path):
                        logd.error(msg)

                yield path

        except StopIteration:
            if (self.cache.parent.meta is not None and
                self.parent.cache != self.cache.anchor and
                self.parent != self):
                # FIXME 99% sure that this never triggers during normal use
                # if you hit this there is a good chance that the path you
                # fed to the DatasetStructure is not a dataset path which
                # will break things
                yield from getattr(self.parent, name_prefix + '_paths')

    @property
    def meta_paths(self):
        for section_name in self.sections:
            glob_type = 'rglob' if section_name in self.rglobs else None
            yield from self._abstracted_paths(section_name, glob_type=glob_type)

    @property
    def data(self):
        # XXX beware if you ever try to call checksum on this it will
        # fail because this is no longer a path in the way that the path
        # checksum exists because this data returns a dict, this could
        # be remedied by coming up with a convention for hashing arbitrary
        # python objects, but we aren't there yet

        out = self.counts
        # out['paths'] = self.inverted_index()
        for section_name in self.sections:
            #section_paths_name = section_name + '_paths'
            #paths = list(getattr(self, section_paths_name))
            glob_type = 'rglob' if section_name in self.rglobs else 'glob'
            paths = list(self._abstracted_paths(section_name, glob_type=glob_type))
            section_key = section_name + '_file'  # XXX _file is added here
            if paths:
                if len(paths) == 1 and section_name != 'manifest':  # FIXME manifest is a hack
                    path, = paths
                    out[section_key] = path
                else:
                    out[section_key] = paths

        out['dir_structure'] = list(self.data_dir_structure().values())

        self.embedErrors(out)
        return out

    xattr_prefix = BlackfynnCache.xattr_prefix
    cache_meta = aug.EatCache.meta

    def data_dir_structure(self):
        #return {d.dataset_relative_path:d.cache_meta for d in self.rchildren_dirs}
        # TODO get all the leaves and then build an inverted index
        def jmc(d): return d.cache._jsonMetadata()
        def jml(d): return d._jsonMetadata()  # XXX TODO FIXME for local validation w/o cache

        def wrap_pathmeta_error(e):
            return [{'message': e,
                     'blame': 'remote',
                     'pipeline_stage': 'DatasetStructure.data_dir_structure',}]
        # FIXME this is just the directories so it is probably ok to
        # resolve the cache for now so we can get the nice json
        # metadata but ultimately we will need _cache_jsonMetadata or similar
        # that uses xattrs to avoid the overhead of constructing the cache class
        # XXX yep, _cache_jsonMetadata is needed especially for standalone validation
        jsonMeta = jmc if self.cache is not None else jml
        instpath = self._cache_class._local_class  # XXXXXXXXXXX FIXME oof inheritance :/
        return {d.dataset_relative_path:{
            **jsonMeta(d),
            'path_meta': {k:wrap_pathmeta_error(v) if k == 'errors' else v
                          for k, v in self.__class__(d).cache_meta.items()
                          # errors can be strings in here! they come from
                          # backends.BlackfynnRemote.meta (for example)
                          if v is not None and (v != tuple()
                                                if k == 'errors'
                                                else True) }}
                for d in (instpath(d) for d in self.rchildren_dirs)}

    def inverted_index(self):
        # TODO there are two steps here
        # 1. identifying invalid folder hierarchies
        #    (e.g. materialized subjects and samples folders)
        # 2. _after_ we have the subjects identifiers we
        #    need to find all the matches remove all the nested
        #    non-matching and then recheck for invalid hierarchies
        # at the moment there seem to be lots of incorrect structures
        # where the primary has a bunch of stuff in between

        # {sourc,primary,derivative}/{subject_id,sample_id,pool_id,performance_id}/{don-t-really-care}...
        dds = self.data_dir_structure()
        #dd = defaultdict(lambda : defaultdict(list))
        dd = defaultdict(list)
        for d in dds:
            p = d.parts
            dd[p[-1]].append((len(p), d, p[::-1]))

        return dict(dd)


class ObjectPath:

    obj = None
    _classes = {}

    _newimpl = DatasetStructure._newimpl

    def __new__(cls, path):
        return cls._newimpl(cls, path)

    @property
    def object(self):
        return self.obj(self)

    def object_new(self, template_schema_version=None):
        """ PROPERTIES ARE BAD KIDS DON'T DO IT UNLESS
            YOU ARE REALLY SURE """  # sigh

        return self.obj(self, template_schema_version)


class Tabular(HasErrors):

    def __new__(cls, *args, **kwargs):
        return super().__new__(cls)

    _renew = __new__

    def __new__(cls, *args, **kwargs):
        # wouldn't it be great if you could bind tabular
        # to _CLASS=Tabular in new and then not have to
        # change the contents of the code and risk difficult
        # to debug copy/paste errors ? LOL PYTHON sigh
        Tabular._setup(*args, **kwargs)
        Tabular.__new__ = Tabular._renew
        return super().__new__(cls)

    @staticmethod
    def _setup(*args, **kwargs):
        import openpyxl
        Tabular._openpyxl = openpyxl

    def __init__(self, path, rotate=False):
        super().__init__()
        self.path = path
        self._rotate = rotate

    @property
    def file_extension(self):
        if self.path.suffixes:
            ext = self.path.suffixes[0]  # FIXME filenames with dots in them ...
            if ext != '.fake':
                return nml.NormFileSuffix(ext).strip('.')

    def tsv(self):
        return self.csv(delimiter='\t')

    def csv(self, delimiter=','):
        # FIXME if the number of delimiters in a row other than the header
        # is greater than the number of delimiters in the header then any
        # columns beyond the number in the header will be truncated and
        # I'm not sure that this loading routine detects that
        for encoding in ('utf-8', 'latin-1'):
            try:
                with open(self.path, 'rt', encoding=encoding) as f:
                    rows_orig = list(csv.reader(f, delimiter=delimiter))

                rows = [row for row in rows_orig if row and any(row)]
                diff = len(rows_orig) - len(rows)
                if diff:
                    # LOL THE FILE WITH > 1 million empty rows, 8mb of commas
                    # totally the maximum errors champion
                    message = f'There are {diff} empty rows in {self.path.as_posix()!r}'
                    if self.addError(message,
                                     blame='submission',
                                     path=self.path):
                        logd.error(message)

                if not rows:
                    ps = self.path.size
                    pcs = self.path.cache.meta.size
                    if ps != pcs:
                        msg = ('We have a likely case of a file that '
                               f'failed to fetch {ps} != {pcs}! {self.path}')
                        log.critical(msg)
                        self.addError(msg,
                                      blame='fetch',
                                      path=self.path)

                    return

                if len(rows[0]) == 1 and ('\t' if delimiter == ',' else ',') in rows[0][0]:
                    message = (f'Possible wrong file extension in {self.path.id} '
                               f'{self.path.project_relative_path}')
                    if self.addError(message,
                                     blame='submission',
                                     path=self.path):
                        logd.error(message)

                if encoding != 'utf-8':
                    message = f'encoding bad {encoding!r} {self.path.as_posix()!r}'
                    if self.addError(exc.EncodingError(message),
                                     blame='submission',
                                     path=self.path):
                        logd.error(message)

                yield from rows
                return
            except UnicodeDecodeError:
                continue
            except csv.Error as e:
                logd.exception(e)
                message = f'WHAT HAVE YOU DONE {e!r} {self.path.as_posix()!r}'
                if self.addError(message,
                                 blame='EVERYONE',
                                 path=self.path):
                    logd.error(message)

    def xlsx(self):
        try:
            one = list(self.xlsx1())
            e1 = None
        except Exception as e:
            e1 = e
            one = None

        try:
            two = list(self.xlsx2())
            e2 = None
        except Exception as e:
            e2 = e
            two = None

        offset = 2
        if one is not None:
            one_test = [['' if c is None else c for c in r[offset:]] for r in one[offset:]]
        else:
            one_test = None

        if two is not None:
            two_test = [['' if c is None else c for c in r[offset:]] for r in two[offset:]]
        else:
            two_test = None

        #if one_test != two_test:
            #breakpoint()

        if e1 is not None:
            message = f'malformed xlsx file could not be read by xlsx2csv {self.path}'
            if self.addError(message,
                             blame='submission',
                             path=self.path):
                logd.exception(e1)

        if e2 is not None:
            message = f'malformed xlsx file could not be read by openpyxl {self.path}'
            if self.addError(message,
                             blame='submission',
                             path=self.path):
                logd.exception(e2)

        if e2 is None:
            yield from two
        elif e1 is None:
            yield from one
        else:
            raise exc.NoDataError(f'{self.path}') from e2

    def _openpyxl_fixes(self):
        # read sheet read only to find empty columns without destroying memory
        wbro = self._openpyxl.load_workbook(self.path, read_only=True)
        if wbro is None:
            raise exc.NoDataError(f'{self.path}')

        lnn = []
        first_cont_empty = 0
        empty_rows = []
        for i, row in enumerate(wbro.active.rows):
            last_not_none = -1
            for j, cell in enumerate(row):
                if isinstance(cell, self._openpyxl.cell.read_only.EmptyCell) or cell.value is None:
                    pass
                else:
                    last_not_none = j
            lnn.append(last_not_none)
            if last_not_none == -1:  # empty row case
                if not empty_rows or empty_rows and empty_rows[-1] < i - 1:
                    # there was an intervening non-empty row
                    first_cont_empty = i

                empty_rows.append(i)

        # non-contiguous empty rows
        sparse_empty_rows = [e for e in empty_rows if e < first_cont_empty]

        wbro.close()
        s_lnn = set(lnn)
        max_lnn = max(s_lnn) + 1  # + 1 to convert from index -> column number
        start_cols_empty = max_lnn + 1  # start deleting one column after last non-empty
        msg = (f'unique last not nones {s_lnn}, max {max_lnn}, '
               f'delete from {start_cols_empty}')
        log.debug(msg)

        # read sheet rw so that hyperlinks and other stuff work
        wb = self._openpyxl.load_workbook(self.path)
        if wb is None:
            raise exc.NoDataError(f'{self.path}')

        wb.iso_dates=True  # XXX NOTE I think only relevant for writing, useful when run in reverse

        sheet = wb.active
        if sheet.max_column > max_lnn:
            # over count a bit using max_column, but safe if max_lnn
            # is somehow zero
            sheet.delete_cols(start_cols_empty, sheet.max_column)
            msg = f'Empty columns beyond {max_lnn} detected in {self.path}'
            if self.addError(msg,
                             blame='submission',
                             path=self.path):
                logd.warning(msg)

        if first_cont_empty > 0:
            # add 1 since row number starts at 1 not 0
            fce_rn = first_cont_empty + 1
            sheet.delete_rows(fce_rn, sheet.max_row)
            msg = f'Empty rows beyond {first_cont_empty} detected in {self.path}'
            if self.addError(msg,
                             blame='submission',
                             path=self.path):
                logd.warning(msg)

        return sheet, wb, sparse_empty_rows

    def xlsx2(self):
        # FIXME this has horrible performance on sheets with lots of empty cells
        # which apparently happens really frequently in our workflows?!
        # https://openpyxl.readthedocs.io/en/latest/datetime.html  # XXX OH NOOOOOOOOOOOOOOOOOOOOOO
        # yeah ... iso8601 as a string :/

        # xlsx: bringing deep sadness to everyone involved
        sheet, wb, _ser = self._openpyxl_fixes()
        longrs = 0
        emptyrs = 0
        for row in sheet.rows:
            fixed_row = [
                cell.value.date()
                # Who thought it was a good idea to make the internal
                # representation of dates and midnight indistinguishable !?!?
                if cell.value and cell.is_date and
                cell.number_format in ('yyyy\\-mm\\-dd;@', 'YYYY\\-MM\\-DD')
                else (cell.hyperlink.target if cell.hyperlink else cell.value)
                for cell in row]
            # NOTE '@' != 'General' but apparently openpyxl converts @ to number incorrectly
            #breakpoint()
            if not [v for v in fixed_row if v]:
                emptyrs += 1
                continue

            yield fixed_row

        wb.close()
        if emptyrs > 0:
            # should only be catching sparse empty rows at this point
            msg = f'{emptyrs} empty rows detected in {self.path}'
            if self.addError(msg,
                             blame='submission',
                             path=self.path):
                logd.warning(msg)

    def xlsx1(self):
        kwargs = {
            'delimiter': '\t',
            'skip_empty_lines': True,
            'outputencoding': 'utf-8',
            'hyperlinks': True,
        }
        sheetid = 1
        try:
            xlsx2csv = Xlsx2csv(self.path.as_posix(), **kwargs)
        except InvalidXlsxFileException as e:
            raise exc.NoDataError(f'{self.path}') from e

        ns = len(xlsx2csv.workbook.sheets)
        if ns > 1:
            message = f'too many sheets ({ns}) in {self.path.as_posix()!r}'
            if self.addError(exc.EncodingError(message),
                             blame='submission',
                             path=self.path):
                logd.error(message)

        f = io.StringIO()
        try:
            xlsx2csv.convert(f, sheetid)
            f.seek(0)
            gen = csv.reader(f, delimiter='\t')
            yield from gen
        except SheetNotFoundException as e:
            log.warning(f'Sheet weirdness in {self.path}')
            log.warning(str(e))
        except AttributeError as e:
            message = ('xlsx2csv: major sheet weirdness ('
                       'probably a bug in the xlsx2csv converter)? '
                       f'in {self.path}')
            if self.addError(exc.EncodingError(message),
                             blame='submission',
                             path=self.path):
                #log.exception(e)
                logd.warning(message)

    def _bad_filetype(self, type_):
        message = f'bad filetype {type_}\n{self.path.as_posix()!r}'
        raise exc.FileTypeError(message)

    def xls(self):
        return self._bad_filetype('xls')

    def normalize(self, rows):
        # FIXME need to detect changes
        # this removes any columns that are all dead

        #if any(not(c) for c in rows[0]):  # doesn't work because generators

        error = exc.EncodingError(f"encoding feff error in '{self.path}'")
        cleaned_rows = zip(*(t for t in zip(*rows) if not all(not(e) for e in t)))  # TODO check perf here
        for row in cleaned_rows:
            n_row = [c.strip().replace('\ufeff', '') if c and isinstance(c, str) else c
                     for c in row
                     if ((not logd.error(error)
                          if self.addError(error) else
                          True)  # FIXME will probably append multiple ...
                         if c and isinstance(c, str) and '\ufeff' in c else
                         True)]
            if not all(not(c) for c in n_row):  # skip totally empty rows
                yield n_row

    @property
    def normalized(self):
        try:
            yield from self.normalize(getattr(self, self.file_extension)())
        except UnicodeDecodeError as e:
            log.error(f'{self.path.as_posix()!r} {e}')

    @property
    def T(self):
        return self.__class__(self.path, rotate=True)

    def __iter__(self):
        yield from self._iter_(normalize=True)

    def _iter_(self, normalize=False):
        try:
            if normalize:
                try:
                    fef = getattr(self, self.file_extension)
                except AttributeError as e:
                    self._bad_filetype(self.file_extension)

                gen = self.normalize(fef())
            else:
                gen = getattr(self, self.file_extension)()

            if self._rotate:
                yield from zip(*gen)
            else:
                yield from gen
        except UnicodeDecodeError as e:
            log.error(f'{self.path.as_posix()!r} {e}')

    @property
    def title(self):
        path = Path(self.path)
        out = path.name
        if path.cache is not None and path.cache.dataset is not None:
            out += ' ' + path.cache.dataset.name[:30] + ' ...'

        return out

    def __write(self, path, format='csv'):
        # FIXME this isn't actually useful because we already have the file in question
        with open(path, 'wt') as f:
            writer = csv.writer(f)
            writer.writerows(self)

    def __repr__(self):
        limit = 30
        rows = [[c[:limit] + ' ...' if isinstance(c, str)
                 and len(c) > limit else c
                 for c in r] for r in self]
        table = AsciiTable(rows, title=self.title)
        return table.table


ROW_TYPE = object()
COLUMN_TYPE = object()
numberN = object()
GROUP_ALL = object()


def remove_rule(blob, rule):
    # FIXME this should be in _DictTransformer
    previous_target = blob
    target = blob
    previous_element = None
    for path_element in rule:
        if isinstance(path_element, str):
            if path_element in target:
                previous_target = target
                target = previous_target[path_element]
            else:
                return  # target not found

        elif isinstance(path_element, dict):
            if all([(k in target and target[k] == v)
                    for k, v in path_element.items()]):
                previous_target.pop(previous_element)
                return  # we're done here

        else:
            raise ValueError(f'what are you doing with {rule}')

        previous_element = path_element


class PrimaryKey:
    def __init__(self, name, column_names, combine_function,
                 nalt_header, agen, gen):
        self.name = name
        # because we add the primary key to alt header we remove it again here
        # because the rest of the generator values don't have a blank prepended
        nalt_header_less_pk = [nah for nah in nalt_header if nah != name]
        try:
            self.indexes = [nalt_header_less_pk.index(name) for name in column_names]
        except ValueError as e:
            msg = f'Missing a primary key component {e}'
            raise exc.MalformedHeaderError(msg) from e

        self.combine_function = combine_function
        self._agen = agen
        self._gen = gen

    def compute_primary(self, arecord):
        return self.combine_function(tuple(arecord[index] for index in self.indexes))

    def __iter__(self):
        yield (self.name, *[self.compute_primary(ar) for ar in self._agen])
        yield from self._gen

    @property
    def generator(self):
        return iter(self)


class MetadataFile(HasErrors):
    default_record_type = ROW_TYPE
    primary_key_rule = None  # name, the names of the alt_header columns to merge, function to merge tuple
    renames_alt = {}     # renames first since all following
    renames_header = {}  # refs look for the renamed form
    missing_add = {}  # missing also comes after renames
    record_type_key_alt = None  # record type key is affected by renames
    record_type_key_header = None  # record type key is affected by renames
    groups_alt = {}
    groups_header = {}
    ignore_alt = tuple()
    ignore_header = tuple()
    ignore_match = tuple()
    raw_json_class = rj.RawJson
    normalization_class = nml.NormValues  # FIXME normalize SHOULD NOT BE DONE AT THIS STAGE
    normalize_alt = True
    normalize_header = True
    normalize_mismatch_ok = tuple()
    normalize_alt_mismatch_ok = tuple()
    _expect_single = tuple()

    def __new__(cls, path, template_schema_version=None):
        if template_schema_version is not None:  # and template_schema_version < latest  # TODO
            logd.debug(f'{template_schema_version} {path}')  # TODO warn

        if cls.record_type_key_header is None or cls.record_type_key_alt is None:
            raise TypeError(f'record_type_key_? should not be None on {cls.__name__}')

        # TODO check for renames that didn't get picked up
        # especially from the 0, 0 cell

        return super().__new__(cls)

    def __init__(self, path, template_schema_version=None):
        super().__init__()
        self.path = path
        self.template_schema_version = template_schema_version

    def xopen(self):
        self.path.xopen()

    @property
    def data(self):
        self._to_embed = tuple()
        data = self._data()

        condition = False
        if condition:
            breakpoint()

        for embedErrors in self._to_embed:
            embedErrors(data)

        self.embedErrors(data)
        return data

    def _data(self):
        if self.path.suffix == '.json':
            data = self.raw_json_class(self.path).data
            self._expand_string_lists = lambda : data  # FIXME hack
            self.norm_to_orig_header = {}
            def all_keys(d):
                if isinstance(d, dict):
                    for k, v in d.items():
                        if k not in self.groups_alt:
                            yield k

                        yield from all_keys(v)

                elif isinstance(d, list):
                    for e in d:
                        yield from all_keys(e)

            self.norm_to_orig_alt = {k:k for k in all_keys(data)}
            #return self.normalization_class(self).data  # FIXME incredible violation of encapsulation ...
            return self._condense()

        else:
            # TODO if this passes we can just render backward from _normalize
            return self._condense()

    def _condense(self):
        def condense(thing, key=None):
            if isinstance(thing, dict):
                hrm = []
                out = {}
                for k, v in thing.items():
                    nv = condense(v, k)
                    # bah bloody description appearing twice >_<
                    if k in self.norm_to_orig_header and k not in self.norm_to_orig_alt:
                        if is_list_or_tuple(nv):
                            # at this stage there are no lists of lists
                            hrm.extend(nv)
                        else:
                            hrm.append(nv)
                    else:
                        out[k] = nv

                if out and hrm:
                    raise ValueError(f'how!?\n{hrm}\n{out}')
                elif out:
                    return out
                elif hrm:
                    if (key in self._expect_single and
                        len(hrm) == 1):
                        return hrm[0]
                    else:
                        return tuple(hrm)

            #elif is_list_or_tuple(thing):
                #return thing
            else:
                return thing

        data_in = self._cull_rtk()
        data_out = condense(data_in)
        if data_out is None:
            # None breaks further pipelines
            # yes types would help here
            return {}
        else:
            return data_out

    def _cull_rtk(self):
        normalized = self._normalize_values()
        def crtk(thing, only_index_alt=False, only_index_header=False):
            if isinstance(thing, dict):
                out = {}
                for k, v in thing.items():
                    if k in self.groups_alt and self.groups_alt[k] == GROUP_ALL:
                        nv = crtk(v, only_index_alt=True)
                        if nv is not None:
                            out[k] = nv
                    elif k in self.groups_header and self.groups_header[k] == GROUP_ALL:
                        # this isn't used yet but putting it here for symmetry just in case
                        nv = crtk(v, only_index_header=True)
                        if nv is not None:
                            out[k] = nv
                    elif (only_index_alt and k == v == self.record_type_key_alt or
                          only_index_header and k == v == self.record_type_key_header):
                        continue
                    elif (k != self.record_type_key_header and
                          # FIXME requires that headers and alts be fully disjoint
                          k != self.record_type_key_alt or
                          only_index_alt and k == self.record_type_key_alt or
                          only_index_header and k == self.record_type_key_header):
                        nv = crtk(v, only_index_alt, only_index_header)
                        if nv is not None:
                            out[k] = nv

                if out:
                    return out

            else:
                return thing

        crv = crtk(normalized)
        return crv

    def _normalize_values(self):
        nc = self.normalization_class(self)  # calls _normalize_keys internally
        return nc.data

    def _clean(self):
        data_in = self._expand_string_lists()
        def clean(thing):
            if isinstance(thing, dict):
                out = {}
                for k, v in thing.items():
                    cv = clean(v)
                    if cv is not None:
                        out[k] = cv

                if out:
                    return out

            elif thing == '' or thing == {}:
                return None

            elif is_list_or_tuple(thing):
                return [v for v in [clean(v) for v in thing] if v is not None]

            else:
                return thing

        data_out = clean(data_in)
        return data_out

    def _expand_string_lists(self):
        data_in = self._filter()
        # TODO consider whether this step would be useful
        # I think it would because it would vastly simplify
        # the downstream implementation of things like
        # normalizing the contributor roles
        # TODO this might also go before filter?
        # might need to blacklist or whitelist fields?
        seps = '|', ';', ','
        data_out = data_in
        return data_out

    def _filter(self):
        # FIXME does in place modification
        culled = self._cull()
        for rule in self.ignore_match:
            remove_rule(culled, rule)

        return culled

    def _cull(self):
        transformed = self._transformed()
        # FIXME deeper nesting might break this?
        culled = {k:{k:v for k, v in v.items()
                     if k not in self.ignore_header}
                  for k, v in transformed.items()}
        return culled

    def _transformed(self):  # FIXME this whole approach makes it very difficult to test individual stages
        keyed = self._keyed()
        naccounted_baseline = set(e for g in self.groups_alt.values()
                                  if g != GROUP_ALL for e in g)

        # TODO how to flag exclude the header fields ?
        # XXX NOTE failure mode on array but should be string? e.g. study_data_collection
        # the order of operations is why this happens, but the behavior is correct
        # fields that are grouped can and do have multiple entires, so if you are
        # expecting a single object where internally there are multipel entries you
        # you do not get that using the group functionality
        transformed = {}
        for key, alt_grouped_norm in self.groups_alt.items():
            merge = alt_grouped_norm == GROUP_ALL
            if merge:
                alt_grouped_norm = tuple(k for k in self.norm_to_orig_alt
                                         if k not in naccounted_baseline)
                naccounted_baseline.update(set(alt_grouped_norm))

            records = [keyed[nh] for nh in alt_grouped_norm
                       # FIXME versioning ... schema will take care of missing values later
                       if nh in keyed]

            objected = {nh:{record[self.record_type_key_header]:record[nh]
                            for record in records}
                        for nh in self.norm_to_orig_header}

            transformed[key] = {k:v for k, v in objected.items()
                                if v is not None and v != {}}

        nunaccounted = (set(self.norm_to_orig_alt) - naccounted_baseline)

        for nah in nunaccounted: 
            transformed[nah] = keyed[nah]

        return transformed

    def _keyed(self):
        objects = self._add_missing()
        keyed = {o[self.record_type_key_header]:o for o in objects}
        return keyed

    def _add_missing(self):
        objects = self._sanity_check()
        for nah, (value, number) in self.missing_add.items():
            if nah not in self.norm_to_orig_alt:
                self.norm_to_orig_alt[nah] = nah
                cant_go_wrong = {}
                i = 0
                for nh in self.norm_to_orig_header:
                    if ((number == numberN or i < number) and
                        nh != self.record_type_key_header and
                        nh not in self.ignore_header):
                        v = value
                        i += 1
                    else:
                        v = ''

                    cant_go_wrong[nh] = v

                {nh:(value if number == numberN or i <= number else '')
                 for i, nh in enumerate(self.norm_to_orig_header)
                 if nh != self.record_type_key_header}
                #print(cant_go_wrong)
                cant_go_wrong[self.record_type_key_header] = nah
                objects.append(cant_go_wrong)

        # end add missing items
        return objects

    def _sanity_check(self):
        """ first chance we have to check sanity since
            the generators have to have run """
        objects = self._objectify()
        for n, rtk, norm_dict in (('alt_header', self.record_type_key_alt, self.norm_to_orig_alt),
                                  ('header', self.record_type_key_header, self.norm_to_orig_header)):
            if rtk not in norm_dict:
                #breakpoint()  # XXX TODO enable this on debug
                msg = (f'Could not find record primary key for {n} in\n"{self.path}"\n\n'
                       f'{rtk}\nnot in\n{list(norm_dict)}')

                raise exc.MalformedHeaderError(msg)

        return objects

    def _objectify(self):
        gen = self._norm_alt_headers()
        # WARNING assumes dict order
        objects = [{k:v for k, v in zip(self.norm_to_orig_header, r)} for r in gen]
        return objects

    def _norm_alt_headers(self):
        # headers MUST be normalized as early as possible due to the fact that
        # there are very likely to be duplicate keys
        try:
            gen = self._headers()
        except exc.BadDataError as e:
            msg = f'Bad data in {self.path}!'
            raise exc.BadDataError(msg) from e

        indexes = {}
        def normb(k):
            if k in self.orig_to_norm_alt:
                nk = self.orig_to_norm_alt[k]
                if isinstance(nk, tuple):
                    if k in indexes:
                        indexes[k] += 1
                    else:
                        indexes[k] = 0

                    return nk[indexes[k]]  # this is safe because _headers is ordered

            else:
                nk = k

            return nk

        for row in gen:
            yield (normb(row[0]), *row[1:])

    def _headers(self):
        def safe_invert(d):
            out = {}
            for v, k in d.items():
                if k in out:
                    ev = out[k]
                    if isinstance(ev, tuple):
                        out[k] += (v,)
                    else:
                        out[k] = (ev, v)
                else:
                    out[k] = v

            return out

        t = self._t()

        if self.default_record_type == ROW_TYPE:
            alt = t
            head = t.T
        else:
            alt = t.T
            head = t

        self._to_embed = alt.embedErrors, head.embedErrors

        agen = iter(alt)
        gen = iter(head)

        try:
            if self.primary_key_rule is not None:
                pk_name, combine_names, combine_function = self.primary_key_rule
                self.alt_header = (pk_name, *next(agen))
            else:
                self.alt_header = next(agen)
        except StopIteration as e:
            raise exc.NoDataError(f'{self.path}') from e

        self._alt_header = Header(self.alt_header,
                                  normalize=self.normalize_alt,
                                  alt=True,
                                  normalize_first_cell=self.normalize_header)
        self.norm_to_orig_alt = self._alt_header.lut(**self.renames_alt)
        self.orig_to_norm_alt = safe_invert(self.norm_to_orig_alt)
        if self.normalize_header and self.alt_header[0] != self._alt_header.data[0]:
            msg = (f'Potentially bad alt header value {self.alt_header[0]!r} '
                   f'!= {self._alt_header.data[0]!r}'
                   f' in {self.path.as_posix()!r}')
            # this is not guranteed to be an issue but hitting
            # Subject_id instead of subject_id will cause issues for
            # other people who are not using this pipeline to
            # normalize everything (for example)
            if self._alt_header.data[0] not in self.normalize_mismatch_ok:
                logd.debug(msg)

        if self.primary_key_rule is not None:
            nalt_header = [self.orig_to_norm_alt[ah] for ah in self.alt_header]
            gen = PrimaryKey(pk_name, combine_names, combine_function,
                             nalt_header, agen, gen).generator

        # FIXME if the 'header' column is not at position zero
        # this will fail see version 1.1 subjects template for this
        # with a note that 1.1 had many issues
        self.header = next(gen)
        self._header = Header(self.header,
                              normalize=self.normalize_header,
                              normalize_first_cell=self.normalize_alt)
        # FIXME the case where someone has a capitalized subject_id
        self.norm_to_orig_header = self._header.lut(**self.renames_header)
        self.orig_to_norm_header = safe_invert(self.norm_to_orig_header)
        if self.normalize_alt and self.header[0] != self._header.data[0]:
            msg = (f'Potentially bad header value {self.header[0]!r} '
                   f'!= {self._header.data[0]!r}'
                   f' in {self.path.as_posix()!r}')
            if self._header.data[0] not in self.normalize_alt_mismatch_ok:
                logd.debug(msg)

        _matched = set(self.norm_to_orig_alt) & set(self.norm_to_orig_header)
        if _matched:
            if (self.record_type_key_header == self.record_type_key_alt and
                {self.record_type_key_header} == _matched):
                # if 0,0 in the sheet is the key for both with the same name
                # then we skip here because it is expected, we might also add
                # a check that it is in ignored, but I don't think we need it
                pass
            else:
                # this has to be fatal because it violates assumptions made by
                # downstream code that are useful to deal with nesting
                # we may be able to fix this in the future but skip for now
                msg = f'Common cells between alt header and header! {_matched}'
                #log.warning(msg)  # can't quite error yet due to metadata_element
                raise exc.MalformedHeaderError(msg)

        yield self.header
        yield from gen

    def _t(self):
        if self.path.suffix == '.json':
            return  # TODO json -> tabular conversion
        else:
            return Tabular(self.path)


class SubmissionFile(MetadataFile):
    __internal_id_1 = object()
    default_record_type = COLUMN_TYPE
    renames_alt = {'submission_item': __internal_id_1}
    renames_header = {'definition': 'definition_header'}
    record_type_key_alt = __internal_id_1
    record_type_key_header = 'submission_item'
    groups_alt = {'submission': ('sparc_award_number', 'milestone_achieved', 'milestone_completion_date')}
    ignore_alt = __internal_id_1,
    ignore_header = 'definition_header',
    raw_json_class = rj.RawJsonSubmission
    normalization_class = nml.NormSubmissionFile

    @property
    def data(self):
        """ lift list with single element to object """

        data = super().data  # LOL PYTHON can't super in a debugger -- pointless
        d = copy.deepcopy(data)

        if d and 'submission' in d:
            sub = d['submission']
            if sub:
                if is_list_or_tuple(sub):
                    d['submission'] = sub[0]
                elif isinstance(sub, dict):
                    return d
                else:
                    raise TypeError(f'What is a {type(sub)} {sub}?')
            else:
                d['submission'] = {}

        return d


class SubmissionFilePath(ObjectPath):
    obj = SubmissionFile


class CodeDescriptionFile(MetadataFile):
    default_record_type = COLUMN_TYPE
    renames_header = {'description': 'description_header'}
    record_type_key_alt = 'metadata_element'
    record_type_key_header = record_type_key_alt
    ignore_header = 'metadata_element', 'description_header', 'example'


class CodeDescriptionFilePath(ObjectPath):
    obj = CodeDescriptionFile


_props = sc.DatasetDescriptionSchema.schema['properties']
_props2 = sc.ContributorSchema.schema['properties']  # FIXME recurse ...
_nddfes = [k for k, v in chain(_props.items(), _props2.items())
           if isinstance(v, dict) and sc.not_array(v)]
_nddfes = sorted(set(_nddfes))


class DatasetDescriptionFile(MetadataFile):
    default_record_type = COLUMN_TYPE
    renames_alt = {'contributors': 'contributor_name',  # watch out for name collisions (heu heu heu)
                   'type': 'dataset_type',  # type by itself is used internally
                   'metadata_version_do_not_change': 'template_schema_version',
                   'contributor_orcid_id': 'contributor_orcid',  # <= 1.2.3
                   'metadata_version': 'template_schema_version',  # 2.0
                   'acknowledgements': 'acknowledgments',  # <= 1.2.3
                   #'relation_type': 'relation_type', # 2.0
                   'identifier': 'related_identifier',
                   'identifier_description': 'related_identifier_description',
                   'identifier_type': 'related_identifier_type',
                   #'study_organ_system': 'organ',  # XXX do not convert, requires normalization
                   #'study_approach': 'approach',  # XXX do not convert, requires normalization
                   #'study_technique': 'technique',  # XXX do not convert, requires normalization
                   }
    renames_header = {'description': 'description_header'}
    missing_add = {'template_schema_version': ('1.0', 1)}
    record_type_key_alt = 'metadata_element'
    record_type_key_header = record_type_key_alt
    groups_alt = {'contributors': ('contributor_name',
                                   'contributor_orcid',
                                   'contributor_affiliation',
                                   'contributor_role',
                                   'is_contact_person',),
                  'related_identifiers': ('related_identifier',
                                          'related_identifier_type',
                                          'relation_type',
                                          'related_identifier_description',),
                  'study': ('study_purpose',
                            'study_data_collection',
                            'study_primary_conclusion',
                            'study_collection_title',),
                  'links': ('additional_links', 'link_description'),
                  'examples': ('example_image_filename',
                               'example_image_locator',
                               'example_image_description'),}

    ignore_header = 'metadata_element', 'example', 'description_header'
    ignore_alt = ('basic_information', 'study_information', 'contributor_information',
                  'related_protocol__paper__dataset__etc_', 'participant_information',)
    raw_json_class = rj.RawJsonDatasetDescription
    normalization_class = nml.NormDatasetDescriptionFile
    normalize_mismatch_ok = 'metadata_element',
    normalize_alt_mismatch_ok = normalize_mismatch_ok
    _expect_single = _nddfes

    @property
    def data(self):
        data = super().data
        return data


class DatasetDescriptionFilePath(ObjectPath):
    obj = DatasetDescriptionFile


_props = sc.SubjectsSchema.schema['properties']['subjects']['items']['properties']
_props2 = sc.SamplesFileSchema.schema['properties']['samples']['items']['properties']
_props3 = sc._software_schema['items']['properties']
_props4 = sc.PerformancesSchema.schema['properties']['performances']['items']['properties']
_nsffes = [k for k, v in chain(_props.items(), _props2.items(), _props3.items(), _props4.items())
           if isinstance(v, dict) and sc.not_array(v)]
_nsffes = sorted(set(_nsffes))


class PerformancesFile(MetadataFile):
    renames_header = {'performance_id': 'metadata_element',}
    record_type_key_alt = 'performance_id'
    record_type_key_header = 'metadata_element'
    groups_alt = {'performances': GROUP_ALL,}
    normalization_class = nml.NormPerformancesFile
    normalize_header = False
    _expect_single = _nsffes


class PerformancesFilePath(ObjectPath):
    obj = PerformancesFile


class SubjectsFile(MetadataFile):
    #default_record_type = COLUMN_TYPE
    __internal_id_1 = object()
    renames_header = {'subject_id': 'metadata_element',
                      ('Lab-based schema for identifying each subject, '
                       'should match folder names'): __internal_id_1,}
    record_type_key_alt = 'subject_id'
    record_type_key_header = 'metadata_element'
    groups_alt = {'subjects': GROUP_ALL,
                  'software':('software',
                              'software_version',
                              'software_vendor',
                              'software_url',
                              'software_rrid'),}
    ignore_header = __internal_id_1,
    ignore_alt = 'additional_fields_e_g__minds',
    ignore_match = [['subjects', 'sub-1', {'subject_id': 'sub-1',
                                           'pool_id': 'pool-1',
                                           'reference_atlas': 'Paxinos Rat V3',
                                           'handedness': 'right',}]]
    raw_json_class = rj.RawJsonSubjects
    normalization_class = nml.NormSubjectsFile
    normalize_header = False
    _expect_single = _nsffes

    @property
    def data(self):
        data = super().data
        return data


class SubjectsFilePath(ObjectPath):
    obj = SubjectsFile


def join_cast(sep):
    """ oh look, you still can't use list comprehension in class scope """
    def sigh(t, sep=sep):
        for _ in t:
            if _ is None:
                raise exc.BadDataError('A primary key value was None!')

        return sep.join(tuple(str(_) for _ in t))

    return sigh


class SamplesFile(SubjectsFile):
    """ TODO ... """

    # TODO merge
    __internal_id_1 = object()
    __primary_key = 'primary_key'
    primary_key_rule = __primary_key, ('subject_id', 'sample_id'), join_cast('_')
    __rename_1 = primary_key_rule[2](('Lab-based schema for identifying each subject',
                                      ('Lab-based schema for identifying each sample, must '
                                       'be unique')))
    renames_header = {__primary_key: 'metadata_element',
                      __rename_1: __internal_id_1,}
    record_type_key_alt = __primary_key
    record_type_key_header = 'metadata_element'
    groups_alt = {'samples': GROUP_ALL,
                  'software':('software',
                              'software_version',
                              'software_vendor',
                              'software_url',
                              'software_rrid'),}
    ignore_header = __internal_id_1,
    ignore_alt = 'additional_fields_e_g__minds',
    ignore_match = [['samples', primary_key_rule[2](('sub-1', 'sub-1_sam-2')),
                     {'subject_id': 'sub-1',
                      'pool_id': 'pool-1',
                      'reference_atlas': 'Paxinos Rat V3',
                      'handedness': 'right',}]]
    normalization_class = nml.NormSamplesFile
    normalize_header = False
    _expect_single = _nsffes


class SamplesFilePath(ObjectPath):
    obj = SamplesFile


# TODO when we need them ?
#_props = sc.ManifestFileSchema.schema['properties']['man']['items']['properties']
#_nsman = [k for k, v in _props.items()
          #if isinstance(v, dict) and sc.not_array(v)]
#_nsman = sorted(set(_nsffes))
class ManifestFile(MetadataFile):  # FIXME need a PatternManifestFile I think?
    renames_header = {'filename': 'metadata_element',}
    record_type_key_alt = 'filename'
    record_type_key_header = 'metadata_element'
    groups_alt = {'manifest_records': GROUP_ALL}
    normalize_header = False
    #raw_json_class = rj.RawJsonManifestFile  # FIXME TODO sigh
    #_expect_single = _nsman

    @staticmethod
    def fix_tod_123(blob):
        if 'timestamp' in blob and 'timestamp_or_date' not in blob:
            # XXX FIXME this destroys the original view of the data
            blob['timestamp_or_date'] = blob.pop('timestamp')

    @property
    def data(self):
        try:
            data = super().data
        except exc.MalformedHeaderError as e:
            if (type(self) == ManifestFile and
                (hasattr(self, 'norm_to_orig_header') and
                 'pattern' in self.norm_to_orig_header and
                 'filename' not in self.norm_to_orig_header or
                 # try to see if pattern is there, but if we have to
                 # run this branch then we are already in trouble
                 'pattern' in self.header and
                 'pattern' in self.alt_header and
                 'filename' not in self.header)):
                log.info(e)  # not much point in logging the full exception
                # FIXME massive hack to work around the multiple
                # different schemas
                return PatternManifestFile(
                    self.path, self.template_schema_version).data
            else:
                raise e

        if isinstance(self.template_schema_version, tuple):
            # error will be caught via schema and will also lead to a
            # multiplication of errors since the elif here won't run
            pass
        elif self.template_schema_version is None:
            # this can happen on a NoDataError
            pass
        elif self.template_schema_version <= '1.2.3': # FIXME HACK
            if 'manifest_records' in data:  # raw json manifests don't conform to this structure FIXME
                [self.fix_tod_123(blob) for blob in data['manifest_records']]

        return data


class PatternManifestFile(ManifestFile):
    # FIXME this is an extreme hack and we need to resolve how we are
    # going to deal with with pattern vs filename because right now
    # you can't really have two valid schemas under a single name,
    # though in principle there is no issue with that
    renames_header = {'pattern': 'metadata_element',}
    record_type_key_alt = 'pattern'
    record_type_key_header = 'metadata_element'
    groups_alt = {'manifest_records': GROUP_ALL}
    normalize_header = False


class ManifestFilePath(ObjectPath):
    obj = ManifestFile


# XXX NOTE *_file is added to all of these keys in DatasetStructure.data
DatasetStructure.sections = {'submission': SubmissionFilePath,
                             'code_description': CodeDescriptionFilePath,
                             'dataset_description': DatasetDescriptionFilePath,
                             'performances': PerformancesFilePath,
                             'subjects': SubjectsFilePath,
                             'samples': SamplesFilePath,
                             'manifest': ManifestFilePath,}

DatasetStructure._bind_sections()
DatasetStructure._pipeline_stage = DatasetStructure.__name__
